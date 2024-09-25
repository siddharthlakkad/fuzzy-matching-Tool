import streamlit as st
import openpyxl
import pandas as pd
from io import BytesIO

# Function to delete rows based on multiple column values
def delete_rows_by_column_values(file, sheet_name, column_letter, values_to_match):
    # Load the workbook from the uploaded file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    
    # Convert the column letter to the index
    column_index = openpyxl.utils.column_index_from_string(column_letter) - 1
    
    rows_to_delete = []
    
    # Collect values in the column for debugging
    column_values = []
    
    # Find all rows matching any of the values
    for row in sheet.iter_rows(min_row=2):  # min_row=2 to skip the header
        cell_value = str(row[column_index].value).strip().lower()  # Cast to string for safe comparison
        column_values.append(cell_value)  # Debugging: collect column values
        
        # Check if the value matches any in the input list
        if cell_value in [v.strip().lower() for v in values_to_match]:
            rows_to_delete.append(row[0].row)
    
    # Delete rows in reverse order to avoid index shifting
    for row_num in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_num)
    
    # Save the modified workbook to a bytes buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return buffer, column_values

# Streamlit UI
st.title("Delete Multiple Rows in Excel Based on Column Data")

# File uploader
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    # Show sheet names
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
    sheet_names = workbook.sheetnames
    sheet_name = st.selectbox("Select sheet", sheet_names)

    # Input for column letter and values to match
    column_letter = st.text_input("Enter column letter (e.g., 'A', 'B')", value="A")
    values_to_match = st.text_area("Enter values to match (comma separated)")

    # Display the Excel file as a preview
    st.write("Uploaded Excel file preview:")
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    st.dataframe(df)

    # Button to trigger deletion
    if st.button("Delete Rows"):
        if column_letter and values_to_match:
            # Convert the input string into a list of values
            values_to_match_list = [v.strip() for v in values_to_match.split(",")]

            # Perform deletion
            updated_file, column_values = delete_rows_by_column_values(uploaded_file, sheet_name, column_letter, values_to_match_list)

            # Show the column values for debugging
            st.write(f"Values in column {column_letter}:")
            st.write(column_values)

            # Provide download link for modified Excel file
            st.success(f"Rows with values {values_to_match_list} in column {column_letter} deleted.")
            st.download_button(label="Download modified Excel file", 
                               data=updated_file, 
                               file_name="modified_file.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.error("Please enter both column letter and values to match.")
